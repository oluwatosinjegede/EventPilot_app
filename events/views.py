import csv
from decimal import Decimal
from io import TextIOWrapper
from django.contrib import messages
from django.core.cache import cache
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.db.models import Count, Q, Sum
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from access.models import DigitalAccessCard
from budgets.models import BudgetItem
from checkins.models import CheckInLog
from contingencies.models import ContingencyPlan
from guests.models import Guest
from invitations.models import GuestInvitation
from invitations.services import send_guest_invitation, validate_qr_token
from logistics.models import LogisticsPlan
from notifications.models import NotificationLog, VendorNotificationRule
from organizations.models import Membership, Organization
from preferences.models import GuestPreference
from promotions.models import PromotionTask
from schedules.models import ScheduleItem
from timelines.models import EventTask
from vendors.models import Vendor
from vendors.models import VendorProfile
from .forms import *
from .models import Event


def home(request): return render(request, 'home.html')

def planner_organization_ids(user):
    if not user.is_authenticated:
        return Membership.objects.none().values('organization_id')
    return Membership.objects.filter(user=user).exclude(role=Membership.VENDOR).values('organization_id')


def user_events(user):
    return (
        Event.objects.filter(organization_id__in=planner_organization_ids(user))
        .select_related('organization')
        .annotate(guest_total=Count('guests', distinct=True))
        .distinct()
    )


def user_organizations(user):
    return Organization.objects.filter(pk__in=planner_organization_ids(user)).distinct()

def get_event_for_user(user, pk): return get_object_or_404(user_events(user), pk=pk)

def event_metrics(event):
    guest_count=event.guests.count(); confirmed=event.guests.filter(rsvp_status='attending').count(); checked=event.guests.filter(checked_in=True).count()
    estimated=event.budget_items.aggregate(v=Sum('estimated_cost'))['v'] or Decimal('0'); actual=event.budget_items.aggregate(v=Sum('actual_cost'))['v'] or Decimal('0')
    return {'guest_count':guest_count,'confirmed_guests':confirmed,'checked_in_guests':checked,'estimated_spend':estimated,'actual_spend':actual,'remaining_budget':event.budget_total-actual,'open_tasks':event.tasks.exclude(status='complete').count(),'vendor_status':event.vendors.values('status').annotate(c=Count('id')),'upcoming_schedule':event.schedule_items.filter(start_time__gte=timezone.now())[:5],'risk_alerts':event.contingency_plans.filter(Q(status__icontains='open')|Q(probability__gte=4)|Q(impact__gte=4))[:5],'cost_per_confirmed_guest': (actual/confirmed if confirmed else 0)}

@login_required
def dashboard(request):
    events = user_events(request.user)
    if VendorProfile.objects.filter(user=request.user, approved=True).exists() and not events.exists():
        return redirect('vendor_dashboard')
    events=user_events(request.user).annotate(guest_total=Count('guests'))[:6]
    return render(request,'events/dashboard.html', {'events':events})

@login_required
def event_list(request): return render(request,'events/list.html', {'events':user_events(request.user).annotate(guest_total=Count('guests'))})

@login_required
def event_create(request):
    orgs=user_organizations(request.user)
    if not orgs.exists(): messages.warning(request,'Create an organization before adding events.'); return redirect('organizations')
    form=EventForm(request.POST or None)
    if request.method=='POST' and form.is_valid():
        event=form.save(commit=False); event.organization=orgs.first(); event.save(); messages.success(request,'Event created.'); return redirect('event_detail', event.pk)
    return render(request,'events/form.html', {'form':form,'title':'Create event'})

@login_required
def event_edit(request, pk):
    event=get_event_for_user(request.user, pk); form=EventForm(request.POST or None, instance=event)
    if request.method=='POST' and form.is_valid(): form.save(); messages.success(request,'Event updated.'); return redirect('event_detail', pk)
    return render(request,'events/form.html', {'form':form,'event':event,'title':'Edit event'})

@login_required
def event_detail(request, pk):
    event=get_event_for_user(request.user, pk)
    return render(request,'events/detail.html', {'event':event, 'metrics':event_metrics(event)})

@login_required
def task_page(request, pk):
    event=get_event_for_user(request.user, pk); form=TaskForm(request.POST or None)
    if request.method=='POST' and form.is_valid(): obj=form.save(commit=False); obj.event=event; obj.save(); return redirect('event_tasks', pk)
    return render(request,'events/module.html', {'event':event,'form':form,'items':event.tasks.all(),'title':'Tasks','kind':'tasks','kanban':EventTask.STATUS_CHOICES})

@login_required
def budget_page(request, pk):
    event=get_event_for_user(request.user, pk); form=BudgetItemForm(request.POST or None)
    if request.method=='POST' and form.is_valid(): obj=form.save(commit=False); obj.event=event; obj.save(); return redirect('event_budget', pk)
    return render(request,'events/module.html', {'event':event,'form':form,'items':event.budget_items.all(),'title':'Budget','kind':'budget','metrics':event_metrics(event)})

@login_required
def vendor_page(request, pk):
    event=get_event_for_user(request.user, pk); vform=VendorForm(request.POST or None, prefix='vendor'); rform=VendorNotificationRuleForm(request.POST or None, prefix='rule')
    if request.method=='POST' and 'save_vendor' in request.POST and vform.is_valid(): obj=vform.save(commit=False); obj.event=event; obj.save(); return redirect('event_vendors', pk)
    if request.method=='POST' and 'save_rule' in request.POST and rform.is_valid(): obj=rform.save(commit=False); obj.event=event; obj.save(); return redirect('event_vendors', pk)
    return render(request,'events/vendors.html', {'event':event,'form':vform,'rule_form':rform,'items':event.vendors.select_related('user'),'rules':event.notification_rules.all(),'title':'Vendors'})

@login_required
def vendor_codes(request, pk):
    event=get_event_for_user(request.user, pk)
    registration_url=request.build_absolute_uri(reverse('vendor_register'))
    return render(request,'vendors/vendor_codes.html', {'event':event,'registration_url':registration_url})

@login_required
def vendor_approve(request, pk, vendor_id):
    event=get_event_for_user(request.user, pk)
    vendor=get_object_or_404(event.vendors, pk=vendor_id)
    VendorProfile.objects.filter(vendor=vendor, event=event).update(approved=True)
    vendor.status='confirmed'; vendor.save(update_fields=['status'])
    messages.success(request, f'{vendor.vendor_name} approved.')
    return redirect('event_vendors', pk)

@login_required
def vendor_reject(request, pk, vendor_id):
    event=get_event_for_user(request.user, pk)
    vendor=get_object_or_404(event.vendors, pk=vendor_id)
    VendorProfile.objects.filter(vendor=vendor, event=event).update(approved=False)
    messages.warning(request, f'{vendor.vendor_name} rejected.')
    return redirect('event_vendors', pk)

@login_required
def guest_page(request, pk):
    event=get_event_for_user(request.user, pk); form=GuestForm(request.POST or None)
    if request.method=='POST' and form.is_valid(): obj=form.save(commit=False); obj.event=event; obj.save(); return redirect('event_guests', pk)
    return render(request,'guests/list.html', {'event':event,'form':form,'guests':event.guests.all()})

def parse_upload(upload):
    name=upload.name.lower(); rows=[]; errors=[]
    if not (name.endswith('.csv') or name.endswith('.xlsx') or name.endswith('.xls')): return [], ['Unsupported file type. Upload CSV or Excel.']
    if name.endswith('.csv'):
        reader=csv.DictReader(TextIOWrapper(upload.file, encoding='utf-8-sig'))
        rows=list(reader)
    else:
        from openpyxl import load_workbook
        wb=load_workbook(upload, read_only=True); ws=wb.active; headers=[str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(max_row=1))]
        for r in ws.iter_rows(min_row=2, values_only=True): rows.append(dict(zip(headers,r)))
    for i,row in enumerate(rows,2):
        if not (row.get('full_name') or row.get('Full name') or row.get('name')): errors.append(f'Row {i}: full_name is required')
    return rows, errors

@login_required
def guest_upload(request, pk):
    event=get_event_for_user(request.user, pk); form=GuestUploadForm(request.POST or None, request.FILES or None); bulk=BulkGuestForm(request.POST or None)
    if request.method=='POST' and 'bulk' in request.POST and bulk.is_valid():
        request.session['guest_import_rows']=[{'full_name':p[0].strip(),'email':p[1].strip() if len(p)>1 else '', 'phone':p[2].strip() if len(p)>2 else ''} for p in (line.split(',') for line in bulk.cleaned_data['guests'].splitlines()) if p and p[0].strip()]
        return redirect('guest_import_preview', pk)
    if request.method=='POST' and 'file' in request.FILES and form.is_valid():
        rows, errors=parse_upload(request.FILES['file']); request.session['guest_import_rows']=rows; request.session['guest_import_errors']=errors; return redirect('guest_import_preview', pk)
    return render(request,'guests/upload.html', {'event':event,'form':form,'bulk':bulk})

@login_required
def guest_import_preview(request, pk):
    event=get_event_for_user(request.user, pk); rows=request.session.get('guest_import_rows', []); errors=request.session.get('guest_import_errors', [])
    if request.method=='POST' and not errors:
        for row in rows:
            Guest.objects.create(event=event, full_name=row.get('full_name') or row.get('Full name') or row.get('name'), email=row.get('email','') or '', phone=row.get('phone','') or '', group_name=row.get('group_name','') or '', access_type=row.get('access_type','General') or 'General', notes=row.get('notes','') or '')
        messages.success(request, f'Imported {len(rows)} guests.'); return redirect('event_guests', pk)
    return render(request,'guests/import_preview.html', {'event':event,'rows':rows,'errors':errors})

def send_invitation(request, guest):
    return send_guest_invitation(guest, request)

@login_required
def send_invites(request, pk):
    event=get_event_for_user(request.user, pk)
    if request.method=='POST':
        sent=0
        for guest in event.guests.all():
            if guest.email: send_invitation(request, guest); sent+=1
        messages.success(request, f'Sent or resent {sent} invitations.'); return redirect('event_guests', pk)
    return render(request,'guests/send_invites.html', {'event':event})

@login_required
def check_in(request, pk):
    event=get_event_for_user(request.user, pk); result=None
    if request.method=='POST':
        ip = request.META.get('REMOTE_ADDR', 'unknown')
        key = f'checkin-rate:{event.pk}:{ip}'
        attempts = cache.get(key, 0)
        if attempts > 120:
            return HttpResponse('Too many check-in attempts. Please wait and try again.', status=429)
        cache.set(key, attempts + 1, 60)
        code=request.POST.get('access_code','').strip()
        result=validate_qr_token(code, event, request.user)
        if result['status'] == 'valid' and result['guest']:
            notify_vendors(event, result['guest'], 'arrived')
    return render(request,'checkins/scanner.html', {'event':event,'result':result,'logs':event.checkin_logs.all()[:20]})

def notify_vendors(event, guest, trigger):
    rules=event.notification_rules.filter(**{f'notify_on_guest_{"confirmed" if trigger=="confirmed" else "arrived"}': True}).select_related('vendor')
    for rule in rules:
        pref=getattr(guest,'preference',None); parts=[f'{guest.full_name} {trigger} for {event.title}.']
        if pref and rule.include_menu: parts.append(f'Menu: {pref.menu_choice}; Allergies: {pref.allergies}')
        if pref and rule.include_seat: parts.append(f'Table/seat: {pref.table_number} {pref.seat_choice}')
        if pref and rule.include_drinks: parts.append(f'Drink: {pref.drink_choice}')
        if pref and rule.include_gift: parts.append(f'Gift: {pref.gift_status} {pref.gift_note}')
        body='\n'.join(parts)
        sent=False
        if rule.vendor.email: send_mail(f'EventPilot update: {guest.full_name}', body, None, [rule.vendor.email]); sent=True
        NotificationLog.objects.create(event=event, vendor=rule.vendor, guest=guest, rule=rule, channel='email' if sent else 'dashboard', subject=f'{guest.full_name} {trigger}', body=body, sent=sent)
        
def simple_module(request, pk, model, formcls, related, title, template='events/module.html'):
    event=get_event_for_user(request.user, pk); instance=None
    if model is LogisticsPlan: instance,_=LogisticsPlan.objects.get_or_create(event=event)
    form=formcls(request.POST or None, request.FILES or None, instance=instance)
    if request.method=='POST' and form.is_valid(): obj=form.save(commit=False); obj.event=event; obj.save(); return redirect(request.path)
    items=[] if instance else getattr(event, related).all()
    return render(request, template, {'event':event,'form':form,'items':items,'title':title,'kind':title.lower()})

@login_required
def logistics_page(request, pk): return simple_module(request, pk, LogisticsPlan, LogisticsPlanForm, 'logistics_plan', 'Logistics')
@login_required
def schedule_page(request, pk): return simple_module(request, pk, ScheduleItem, ScheduleItemForm, 'schedule_items', 'Schedule')
@login_required
def promotions_page(request, pk): return simple_module(request, pk, PromotionTask, PromotionTaskForm, 'promotion_tasks', 'Promotions')
@login_required
def contingencies_page(request, pk): return simple_module(request, pk, ContingencyPlan, ContingencyPlanForm, 'contingency_plans', 'Contingencies')
@login_required
def reports_page(request, pk):
    event=get_event_for_user(request.user, pk)
    return render(request,'events/reports.html', {'event':event,'metrics':event_metrics(event),'notifications':event.notification_logs.all()[:25]})
